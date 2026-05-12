#!/usr/bin/env python3
"""
CLLS出欠一覧.xlsx の「参加対象一覧」シートを読み取り、正規化したCSVを生成する。
列: グループ番号,グループ名,クラブ名,名前,フリガナ,役職,大タグ
"""
import sys
import re
import csv
import openpyxl
from pathlib import Path

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else '/Users/furuken/Downloads/CLLS出欠一覧 (3).xlsx'
OUT_CSV = sys.argv[2] if len(sys.argv) > 2 else 'data/members_real.csv'
MAP_REPORT = 'data/club_mapping.tsv'
UNMATCHED_REPORT = 'data/unmatched.tsv'

# ───────────────────────────────────────────────
# グループ定義 (1〜7 = 地理的グループ, 8 = ガバナー事務所)
GROUPS = {
    1: '第1グループ',
    2: '第2グループ',
    3: '第3グループ',
    4: '第4グループ',
    5: '第5グループ',
    6: '第6グループ',
    7: '第7グループ',
    8: 'ガバナー事務所',
}

# クラブ → グループ番号 のマスタ (短形ベース、suffix なしのカノニカル名)
# RACは ローターアクト系。 RC suffix で統一。
CLUB_GROUP = {
    # 第1グループ
    '豊前RC': 1, '豊前西RC': 1, '苅田RC': 1, '田川RC': 1, '行橋RC': 1, '行橋COSMOSRSC': 1, '行橋みやこRC': 1,
    # 第2グループ
    '小倉RC': 2, '小倉中央RC': 2, '小倉東RC': 2, '小倉南RC': 2, '小倉西RC': 2,
    '門司RC': 2, '門司西RC': 2, '門司西めかりRC': 2,
    '戸畑RC': 2, '戸畑東RC': 2, '若松RC': 2, '若松中央RC': 2,
    # 第3グループ
    '飯塚RC': 3, '直方RC': 3, '直方中央RC': 3, '遠賀RC': 3,
    '八幡RC': 3, '八幡中央RC': 3, '八幡南RC': 3, '八幡西RC': 3, '八幡RAC': 3,
    # 第4グループ
    '太宰府RC': 4, '福岡RC': 4, '福岡エアポートRC': 4, '福岡平成RC': 4,
    '福岡東RC': 4, '福岡東令和あけぼのRSC': 4, '福岡城南RC': 4, '福岡南RC': 4,
    '福岡南ファミリアRSC': 4, '福岡南RAC': 4, '福岡RAC': 4,
    '福岡東南RC': 4, '福岡東南けやきRSC': 4,
    '博多イブニングRC': 4, '博多イブニングトワイライトRSC': 4,
    '宗像RC': 4, '対馬RC': 4, '対馬ちんぐRSC': 4,
    # 第5グループ
    '福岡中央RC': 5, '福岡中央エンジョイRSC': 5, '福岡中央RAC': 5,
    '福岡イブニングRC': 5, '福岡城西RC': 5, '福岡城東RC': 5,
    '福岡北RC': 5, '福岡西RC': 5,
    '博多RC': 5, '壱岐RC': 5, '壱岐中央RC': 5, '糸島RC': 5,
    # 第6グループ
    '甘木RC': 6, '久留米RC': 6, '久留米中央RC': 6, '久留米中央みらいRSC': 6,
    '久留米東RC': 6, '久留米北RC': 6,
    '小郡RC': 6, '小郡七夕RSC': 6, '鳥栖RC': 6, '浮羽RC': 6,
    # 第7グループ
    '筑後RC': 7, '大川RC': 7,
    '大牟田RC': 7, '大牟田北RC': 7, '大牟田南RC': 7,
    '八女RC': 7, '八女グリーンRSC': 7, '柳川RC': 7,
    # 第8グループ (ガバナー事務所)
    'ガバナー事務所': 8,
}

# ───────────────────────────────────────────────
# 正規化ルール
def normalize_club(raw: str) -> str:
    """Excelに書かれたクラブ名表記をカノニカル名に揃える"""
    if not raw:
        return ''
    s = raw.strip()
    # 改行・全角空白・半角空白・zero-width space を除去
    s = s.replace('\n', '').replace('\r', '')
    s = s.replace('　', '').replace(' ', '').replace('​', '')

    # 別名・別表記の救済
    aliases = {
        # 全角space入り
        '飯塚': '飯塚RC',
        '柳川': '柳川RC',
        # サテライト・ローターアクトの長形 → suffix付きカノニカル
    }
    if s in aliases:
        return aliases[s]

    # 既にカノニカル形式 (RSC / RAC / RC で終わる) ならそのまま
    if s.endswith('RSC') or s.endswith('RAC'):
        return s
    if s.endswith('RC') and not s.endswith('ロータリークラブ'):
        return s
    # サフィックス処理
    # "○○ロータリー衛星クラブ" → "○○RSC"
    if s.endswith('ロータリー衛星クラブ'):
        return s[: -len('ロータリー衛星クラブ')] + 'RSC'
    # "○○ローターアクトクラブ" or "○○ローターアクト" → "○○RAC"
    if s.endswith('ローターアクトクラブ'):
        return s[: -len('ローターアクトクラブ')] + 'RAC'
    if s.endswith('ローターアクト'):
        return s[: -len('ローターアクト')] + 'RAC'
    # "○○ロータリークラブ" → "○○RC"
    if s.endswith('ロータリークラブ'):
        return s[: -len('ロータリークラブ')] + 'RC'
    # ガバナー事務所はそのまま
    if s == 'ガバナー事務所':
        return s
    # 短形（豊前 など）→ "○○RC"
    return s + 'RC'


# ───────────────────────────────────────────────
# 大タグ判定（カノニカルな短い名前に変換）
TAG_CANON = {
    '■2026-27年度・三役':              '2026-27年度・三役',
    '■ゲスト・講演者':                 'ゲスト・講演者',
    '■2026-27年度・ガバナー補佐':      '2026-27年度・ガバナー補佐',
    '■協議会主導 委員長':              '協議会主導・委員長',
    '■協議会 講演者(会長、幹事部門)':   None,  # 後で置換 - placeholder
    '■支援室':                         '支援室',
}
# タグの優先順位 (重複時にどちらを採用するか)
TAG_PRIORITY = {
    '2026-27年度・三役': 1,
    'ゲスト・講演者': 2,
    '2026-27年度・ガバナー補佐': 3,
    '協議会主導・委員長': 4,
    '支援室': 5,
    '': 99,  # 一般会員 (最下位)
}

# 括弧違いの ASCII vs 全角
TAG_LOOKUP = {}
for k, v in TAG_CANON.items():
    TAG_LOOKUP[k] = v
# 協議会講演者セクションの中で「◆各クラブ出席者」以前にいる委員長は「協議会主導・委員長」と統合
# (R36-R37 の RLI委員長 / 広報・公共イメージ委員長 など)
# ◆各クラブ出席者 以降は空タグ (一般会員扱い)
TAG_LOOKUP['■協議会 講演者（会長、幹事部門）'] = '協議会主導・委員長'
TAG_LOOKUP['■協議会 講演者(会長、幹事部門)'] = '協議会主導・委員長'


# ───────────────────────────────────────────────
def parse_committee_sheet(wb):
    """シート '協議会別参加者' を解析し、(name, club_canonical) → (committee, room) の dict を返す。
    name は (対面)/(オンライン)/(代理) などの註釈を除去した形に正規化する。"""
    if '協議会別参加者' not in wb.sheetnames:
        return {}
    ws = wb['協議会別参加者']

    # R2 がクラブ短名のヘッダー (C8〜C83)
    col_to_club = {}
    for c in range(8, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v and isinstance(v, str):
            col_to_club[c] = normalize_club(v)

    member_to_committee = {}
    current_committee = None
    current_room = None

    for r in range(3, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        c5 = ws.cell(row=r, column=5).value

        # 主行 (C1 に数字) → 新しい委員会の開始
        if c1 is not None and isinstance(c1, (int, float)):
            current_committee = c2.strip() if isinstance(c2, str) else None
            current_room = c5.strip() if isinstance(c5, str) else None

        if not current_committee:
            continue

        # 「参加人数」など、メンバー欄ではない総計行を識別 (C5 が '参加人数' など)
        if isinstance(c5, str) and c5 == '参加人数':
            continue

        for c, club in col_to_club.items():
            v = ws.cell(row=r, column=c).value
            if not v or not isinstance(v, str):
                continue
            # 註釈を除去: (対面) / （対面） / (オンライン) / （代理） など
            name = re.sub(r'[(（].*?[)）]', '', v).replace('　', ' ').strip()
            # 連続スペースを1つに
            name = re.sub(r'\s+', ' ', name)
            if not name:
                continue
            # 数字や非名前は除外
            if name.isdigit():
                continue
            # 備考行を除外 (20文字以上 or 「様」終わり or 句点を含む)
            if len(name) > 18 or name.endswith('様') or '。' in name or '：' in name:
                continue
            member_to_committee[(name, club)] = (current_committee, current_room)

    return member_to_committee


def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['参加対象一覧']

    rows = []
    current_tag = None
    sub_header = None  # ☆第Nグループ や 当日運営補助
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value  # 役職
        c3 = ws.cell(row=r, column=3).value  # 名前
        c4 = ws.cell(row=r, column=4).value  # かな

        # ■大タグ
        if isinstance(c1, str) and c1.startswith('■'):
            current_tag = TAG_LOOKUP.get(c1, c1.lstrip('■'))
            sub_header = None
            continue
        # ◆/☆ サブ見出しと「当日運営補助」
        if isinstance(c1, str) and (c1.startswith('◆') or c1.startswith('☆') or c1 == '当日運営補助'):
            sub_header = c1
            # 「協議会講演者」セクション内で◆各クラブ出席者が出たら、それ以降は空タグ (一般会員)
            if c1.startswith('◆') and current_tag == '協議会主導・委員長':
                current_tag = ''
            continue

        # 会員行: 名前(C3)があるなら採用
        if not c3:
            continue
        # まだ ■大タグが現れる前のヘッダー行(R1-R3 など)はスキップ
        if current_tag is None:
            continue
        name = re.sub(r'\s+', ' ', str(c3).replace('　', ' ').strip())
        kana = (str(c4) if c4 else '').replace('　', ' ').strip()
        role = (str(c2) if c2 else '').replace('　', ' ').strip()
        raw_club = str(c1) if c1 else ''
        club = normalize_club(raw_club) if raw_club else 'ガバナー事務所'

        rows.append({
            'raw_club': raw_club,
            'club': club,
            'name': name,
            'kana': kana,
            'role': role,
            'tag': current_tag or '',
            'sub_header': sub_header or '',
            'src_row': r,
        })
    return rows


def main():
    out_path = Path(OUT_CSV)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # 参加対象一覧を読む
    rows = parse_excel(XLSX_PATH)

    # 協議会別参加者シートを読み、メンバーに委員会情報を紐付ける
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    committee_map = parse_committee_sheet(wb)
    for row in rows:
        key = (row['name'], row['club'])
        committee, room = committee_map.get(key, ('', ''))
        row['committee'] = committee
        row['committee_room'] = room

    # 参加対象一覧にないが協議会別参加者シートに居る人を補完追加
    existing_keys = set((r['name'], r['club']) for r in rows)
    appended_from_committee = 0
    for (name, club), (committee, room) in committee_map.items():
        if (name, club) in existing_keys:
            continue
        if club not in CLUB_GROUP:
            continue  # 未知のクラブはスキップ (基本ない)
        rows.append({
            'raw_club': '',
            'club': club,
            'name': name,
            'kana': '',     # フリガナ後で追加
            'role': '',     # 役職不明
            'tag': '',
            'sub_header': '(協議会から補完)',
            'src_row': 0,
            'committee': committee,
            'committee_room': room,
        })
        appended_from_committee += 1
    if appended_from_committee:
        print(f'協議会シートから補完追加: {appended_from_committee}名')

    # 委員会マッチ統計
    committee_assigned = sum(1 for r in rows if r.get('committee'))
    print(f'委員会紐付け: {committee_assigned} / {len(rows)} 人')
    unmatched_in_committee = []
    used_keys = set()
    for row in rows:
        used_keys.add((row['name'], row['club']))
    for key in committee_map:
        if key not in used_keys:
            unmatched_in_committee.append(key)
    if unmatched_in_committee:
        print(f'  協議会シートにあるが参加対象一覧と一致しないメンバー: {len(unmatched_in_committee)}件')
        with open('data/committee_unmatched.tsv', 'w', encoding='utf-8') as f:
            f.write('名前\tクラブ\n')
            for n, c in unmatched_in_committee:
                f.write(f'{n}\t{c}\n')
        print(f'  -> data/committee_unmatched.tsv に書き出し')

    # ── 重複排除: 同一(名前, クラブ) はタグ優先順位の高い1件にまとめる ──
    by_key = {}
    dup_log = []
    for row in rows:
        key = (row['name'], row['club'])
        prio = TAG_PRIORITY.get(row['tag'], 50)
        if key not in by_key:
            by_key[key] = (prio, row)
        else:
            old_prio, old_row = by_key[key]
            if prio < old_prio:
                # 新しい方を採用、古い方をログ
                dup_log.append((row['name'], row['club'], old_row['tag'] or '会員', row['tag'] or '会員', '採用=' + (row['tag'] or '会員')))
                by_key[key] = (prio, row)
            else:
                dup_log.append((row['name'], row['club'], row['tag'] or '会員', old_row['tag'] or '会員', '採用=' + (old_row['tag'] or '会員')))
    rows = [v[1] for v in by_key.values()]

    if dup_log:
        with open('data/duplicates_log.tsv', 'w', encoding='utf-8') as f:
            f.write('名前\tクラブ\t片方のタグ\tもう片方のタグ\t採用\n')
            for d in dup_log:
                f.write('\t'.join(d) + '\n')

    # 検証 & マッピング集計
    unique_clubs = {}
    unmatched = []  # (raw, normalized)
    for row in rows:
        norm = row['club']
        unique_clubs.setdefault(norm, {'raw_forms': set(), 'count': 0})
        unique_clubs[norm]['raw_forms'].add(row['raw_club'] or '(空欄)')
        unique_clubs[norm]['count'] += 1
        if norm not in CLUB_GROUP:
            unmatched.append((row['raw_club'], norm, row['src_row']))

    # マッピング表
    with open(MAP_REPORT, 'w', encoding='utf-8') as f:
        f.write('正規化クラブ名\tグループ\t件数\tExcel原文\n')
        for norm in sorted(unique_clubs.keys()):
            gid = CLUB_GROUP.get(norm, '?')
            gname = GROUPS.get(gid, '?') if isinstance(gid, int) else '?'
            forms = ' / '.join(sorted(unique_clubs[norm]['raw_forms']))
            f.write(f'{norm}\t{gid} ({gname})\t{unique_clubs[norm]["count"]}\t{forms}\n')

    with open(UNMATCHED_REPORT, 'w', encoding='utf-8') as f:
        f.write('Excel原文\t正規化後\t行番号\n')
        for raw, norm, src in unmatched:
            f.write(f'{raw}\t{norm}\t{src}\n')

    # CSV出力（インポートAPIに食わせる形式: 9列）
    with open(out_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['グループ番号', 'グループ名', 'クラブ名', '名前', 'フリガナ', '役職', '大タグ', '協議会', '部屋'])
        skipped = 0
        for row in rows:
            gid = CLUB_GROUP.get(row['club'])
            if gid is None:
                skipped += 1
                continue
            w.writerow([
                gid,
                GROUPS[gid],
                row['club'],
                row['name'],
                row['kana'],
                row['role'],
                row['tag'],
                row.get('committee', ''),
                row.get('committee_room', ''),
            ])

    # サマリ
    tag_count = {}
    for row in rows:
        tag_count[row['tag'] or '(なし)'] = tag_count.get(row['tag'] or '(なし)', 0) + 1

    print('=== サマリ ===')
    print(f'  Excel行数: {len(rows)}')
    print(f'  ユニーククラブ: {len(unique_clubs)}')
    print(f'  未マッチクラブ: {len(set((u[1] for u in unmatched)))}件')
    if unmatched:
        print(f'    -> {UNMATCHED_REPORT} を確認')
    print(f'  CSV出力: {out_path} (skipped={skipped})')
    print(f'  マッピング表: {MAP_REPORT}')
    print()
    print('=== 大タグ別件数 ===')
    for t, c in sorted(tag_count.items(), key=lambda x: -x[1]):
        print(f'  {t}: {c}人')


if __name__ == '__main__':
    main()
